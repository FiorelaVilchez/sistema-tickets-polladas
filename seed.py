import sys
import os
import openpyxl

# Agregar directorio actual al sys.path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from datetime import datetime
from app.database import engine, SessionLocal, Base
from app.models import Usuario, Evento, Ticket, EstudianteMatriculado
from app.auth import get_password_hash

def seed_database():
    print("Recreando tablas en la base de datos con la nueva estructura...")
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    try:
        # 1. Cargar Estudiantes Matriculados desde EstudiantesMatriculados.xlsx
        excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "EstudiantesMatriculados.xlsx")
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            estudiantes_agregados = 0
            for r in range(7, ws.max_row + 1):
                val_codigo = ws.cell(r, 1).value
                val_nombre = ws.cell(r, 2).value
                val_carrera = ws.cell(r, 3).value
                val_ciclo = ws.cell(r, 4).value

                if val_codigo and val_nombre:
                    codigo_str = str(val_codigo).strip()
                    nombre_str = str(val_nombre).strip()
                    carrera_str = str(val_carrera).strip() if val_carrera else "INGENIERIA DE SISTEMAS"
                    ciclo_str = str(val_ciclo).strip() if val_ciclo else "1"

                    estudiante = EstudianteMatriculado(
                        codigo=codigo_str,
                        nombre=nombre_str,
                        carrera=carrera_str,
                        ciclo=ciclo_str
                    )
                    db.add(estudiante)
                    estudiantes_agregados += 1

            db.commit()
            print(f"[OK] {estudiantes_agregados} estudiantes matriculados cargados desde Excel.")
        else:
            print("[WARN] No se encontro el archivo EstudiantesMatriculados.xlsx")

        # 2. Crear usuario admin por defecto
        admin_user = Usuario(
            username="admin",
            password_hash=get_password_hash("admin123")
        )
        db.add(admin_user)
        print("[OK] Usuario por defecto creado -> Username: admin | Password: admin123")

        # 3. Crear eventos de prueba
        evento1 = Evento(
            nombre="Pollada 2026",
            fecha=datetime(2026, 10, 15, 12, 0, 0),
            activo=True
        )
        evento2 = Evento(
            nombre="Cachimbeada 2026",
            fecha=datetime(2026, 11, 20, 20, 0, 0),
            activo=True
        )
        db.add_all([evento1, evento2])
        db.commit()
        print("[OK] Eventos iniciales creados: 'Pollada 2026' y 'Cachimbeada 2026'")

        db.commit()
        print("\nBase de datos inicializada correctamente.")
    except Exception as e:
        db.rollback()
        print(f"[ERROR] Error al inicializar la base de datos: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    seed_database()
