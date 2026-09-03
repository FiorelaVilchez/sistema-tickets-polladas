import os
import sys
from fastapi.testclient import TestClient
import openpyxl
import io

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app.main import app
from seed import seed_database

def run_tests():
    print("=== 1. Ejecutando Seed para cargar EstudiantesMatriculados.xlsx ===")
    seed_database()
    client = TestClient(app)

    print("\n=== 2. Login para obtener Token JWT ===")
    resp_login = client.post("/auth/login", data={"username": "admin", "password": "admin123"})
    assert resp_login.status_code == 200
    token = resp_login.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Obtener ID de un evento existente
    resp_eventos = client.get("/eventos", headers=headers)
    assert resp_eventos.status_code == 200
    evento_id = resp_eventos.json()[0]["id"]

    print("\n--- TEST 3: Probar Autocompletado de Estudiante Matriculado ---")
    r_est = client.get("/estudiantes/buscar?q=247358", headers=headers)
    assert r_est.status_code == 200
    estudiantes = r_est.json()
    assert len(estudiantes) > 0
    assert estudiantes[0]["codigo"] == "247358"
    assert "AFARAYA" in estudiantes[0]["nombre"]
    print(f"[OK] Estudiante encontrado: {estudiantes[0]['codigo']} - {estudiantes[0]['nombre']} ({estudiantes[0]['carrera']})")

    print("\n--- TEST 4: Registrar Venta Múltiple con Boletos Numerados y Recolectores Diferentes ---")
    venta_payload = {
        "evento_id": evento_id,
        "codigo_alumno": "247358",
        "nombre_alumno": "AFARAYA HUANCCO JHON RICARDO",
        "carrera": "INGENIERIA DE SISTEMAS",
        "ciclo": "5",
        "estado": "parcialmente_pagado",
        "precio_unitario": 15.0,
        "monto_pagado_total": 30.0, # Pagó S/ 30.00 de S/ 45.00 total
        "metodo_pago": "yape",
        "boletos": [
            {"numero_boleto": 101, "nombre_recolector": "Pepe"},
            {"numero_boleto": 102, "nombre_recolector": "María"},
            {"numero_boleto": 103, "nombre_recolector": "Jhon Ricardo"}
        ]
    }

    r_venta = client.post("/tickets/registrar-venta", json=venta_payload, headers=headers)
    assert r_venta.status_code == 201, f"Error al registrar venta: {r_venta.text}"
    tickets = r_venta.json()
    assert len(tickets) == 3
    assert tickets[0]["numero_boleto"] == 101
    assert tickets[0]["nombre_recolector"] == "Pepe"
    assert tickets[1]["numero_boleto"] == 102
    assert tickets[1]["nombre_recolector"] == "María"
    print(f"[OK] Venta múltiple registrada correctamente para {len(tickets)} boletos físicos (#101, #102, #103).")

    print("\n--- TEST 5: Validar Rechazo de Boleto Físico Duplicado ---")
    venta_duplicada = {
        "evento_id": evento_id,
        "codigo_alumno": "202499",
        "nombre_alumno": "Alumno Prueba",
        "carrera": "Inteligencia Artificial",
        "ciclo": "2",
        "estado": "pagado",
        "precio_unitario": 15.0,
        "monto_pagado_total": 15.0,
        "metodo_pago": "efectivo",
        "boletos": [
            {"numero_boleto": 102, "nombre_recolector": "Prueba"}
        ]
    }
    r_dup = client.post("/tickets/registrar-venta", json=venta_duplicada, headers=headers)
    assert r_dup.status_code == 400
    assert "ya han sido vendidos o asignados" in r_dup.json()["detail"]
    print(f"[OK] Intento de duplicar boleto físico #102 capturado correctamente: '{r_dup.json()['detail']}'")

    print("\n--- TEST 6: Búsqueda por Número de Boleto, Código o Persona que Recoge ---")
    rb1 = client.get("/tickets/buscar?q=Pepe", headers=headers)
    assert rb1.status_code == 200
    assert len(rb1.json()) == 1
    assert rb1.json()[0]["numero_boleto"] == 101
    print(f"[OK] Búsqueda por recolector 'Pepe' devolvió Boleto #{rb1.json()[0]['numero_boleto']}.")

    print("\n--- TEST 7: Confirmar Entrega y Saldo Pendiente ---")
    entrega_payload = {
        "numero_boleto": 101,
        "monto_cobrado_adicional": 5.0, # Cobrar saldo en puerta
        "metodo_pago_entrega": "efectivo"
    }
    r_ent = client.post("/tickets/confirmar-entrega", json=entrega_payload, headers=headers)
    assert r_ent.status_code == 200
    ticket_entregado = r_ent.json()
    assert ticket_entregado["entregado"] is True
    assert ticket_entregado["fecha_hora_entrega"] is not None
    print(f"[OK] Entrega del boleto #101 confirmada a las {ticket_entregado['fecha_hora_entrega']}.")

    # Intento de entrega duplicada
    r_ent_dup = client.post("/tickets/confirmar-entrega", json={"numero_boleto": 101}, headers=headers)
    assert r_ent_dup.status_code == 400
    assert "ya fue entregado previamente" in r_ent_dup.json()["detail"]
    print(f"[OK] Bloqueo de entrega duplicada verificado: '{r_ent_dup.json()['detail']}'")

    print("\n--- TEST 8: KPIs con Métricas de Cobertura de Matriculados y Recaudación ---")
    rk = client.get(f"/eventos/{evento_id}/kpis", headers=headers)
    assert rk.status_code == 200
    kpis = rk.json()
    print(f"[OK] KPIs del evento '{kpis['nombre_evento']}':")
    print(f"     Estudiantes Matriculados Con Boleto: {kpis['estudiantes_matriculados_con_boleto']} / {kpis['estudiantes_matriculados_total']} ({kpis['porcentaje_cobertura_matriculados']}%)")
    print(f"     Total Boletos: {kpis['total_boletos']}")
    print(f"     Total Recaudado: S/ {kpis['total_recaudado']:.2f}")
    print(f"     Saldo Pendiente: S/ {kpis['total_pendiente']:.2f}")
    assert kpis["estudiantes_matriculados_con_boleto"] >= 1

    print("\n--- TEST 9: Exportación Excel y CSV ---")
    re_excel = client.get(f"/eventos/{evento_id}/exportar/excel", headers=headers)
    assert re_excel.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(re_excel.content))
    ws = wb.active
    assert ws.max_row >= 4
    print(f"[OK] Archivo Excel generado ({len(re_excel.content)} bytes, {ws.max_row} filas).")

    re_csv = client.get(f"/eventos/{evento_id}/exportar/csv", headers=headers)
    assert re_csv.status_code == 200
    assert "Nº Boleto" in re_csv.content.decode('utf-8')
    print(f"[OK] Archivo CSV generado ({len(re_csv.content)} bytes).")

    print("\n=======================================================================")
    print("  TODAS LAS PRUEBAS DE LA NUEVA ESTRUCTURA DE BOLETOS PASARON CON ÉXITO")
    print("=======================================================================")

if __name__ == "__main__":
    run_tests()
