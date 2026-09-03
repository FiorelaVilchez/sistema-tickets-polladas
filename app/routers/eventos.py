import csv
import io
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Response, status
from sqlalchemy.orm import Session
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models import Evento, Ticket, EstudianteMatriculado, Usuario
from app.schemas import EventoCreate, EventoUpdate, EventoOut, EventoKPIsOut
from app.auth import get_current_user

router = APIRouter(prefix="/eventos", tags=["Eventos y Reportes"])

def calcular_kpis_evento(evento: Evento, db: Session) -> EventoKPIsOut:
    tickets = db.query(Ticket).filter(Ticket.evento_id == evento.id).all()
    
    total_boletos = len(tickets)
    pagados = sum(1 for t in tickets if t.estado == "pagado")
    parcialmente_pagados = sum(1 for t in tickets if t.estado == "parcialmente_pagado")
    separados = sum(1 for t in tickets if t.estado == "separado")
    entregados = sum(1 for t in tickets if t.entregado)

    total_recaudado = sum(t.monto_pagado for t in tickets)
    total_pendiente = sum(t.monto_pendiente for t in tickets)

    # Cobertura de estudiantes matriculados (EstudiantesMatriculados.xlsx)
    total_matriculados = db.query(EstudianteMatriculado).count()
    
    # Obtener códigos únicos de compradores en este evento que son estudiantes matriculados
    codigos_compradores = db.query(Ticket.codigo_alumno).filter(Ticket.evento_id == evento.id).distinct().all()
    set_codigos_compradores = set(c[0] for c in codigos_compradores)
    
    matriculados_con_boleto = db.query(EstudianteMatriculado).filter(
        EstudianteMatriculado.codigo.in_(list(set_codigos_compradores))
    ).count() if set_codigos_compradores else 0

    porcentaje_cobertura = round((matriculados_con_boleto / total_matriculados) * 100, 2) if total_matriculados > 0 else 0.0

    return EventoKPIsOut(
        evento_id=evento.id,
        nombre_evento=evento.nombre,
        total_boletos=total_boletos,
        pagados=pagados,
        parcialmente_pagados=parcialmente_pagados,
        separados=separados,
        entregados=entregados,
        total_recaudado=round(total_recaudado, 2),
        total_pendiente=round(total_pendiente, 2),
        estudiantes_matriculados_total=total_matriculados,
        estudiantes_matriculados_con_boleto=matriculados_con_boleto,
        porcentaje_cobertura_matriculados=porcentaje_cobertura
    )

@router.get("/kpis", response_model=List[EventoKPIsOut])
def obtener_kpis_todos_eventos(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Devuelve los KPIs consolidados de todos los eventos.
    Requiere autenticación.
    """
    eventos = db.query(Evento).all()
    return [calcular_kpis_evento(e, db) for e in eventos]

@router.get("/{evento_id}/kpis", response_model=EventoKPIsOut)
def obtener_kpis_evento(
    evento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Devuelve las métricas KPI y porcentaje de cobertura de estudiantes matriculados para un evento específico.
    Requiere autenticación.
    """
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento no encontrado"
        )
    return calcular_kpis_evento(evento, db)

@router.get("/{evento_id}/exportar/excel")
def exportar_tickets_excel(
    evento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Exporta la lista completa de boletos vendidos/separados a un archivo Excel (.xlsx) con formato profesional.
    Requiere autenticación.
    """
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento no encontrado"
        )

    tickets = db.query(Ticket).filter(Ticket.evento_id == evento_id).order_by(Ticket.numero_boleto.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Boletos"

    # Estilos Excel
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = [
        "Nº Boleto", "Código / DNI", "Nombre Comprador", "Carrera", "Ciclo",
        "Persona que Recoge", "Estado", "Monto Total (S/)", "Monto Pagado (S/)",
        "Monto Pendiente (S/)", "Método de Pago", "Entregado", "Fecha de Entrega"
    ]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for t in tickets:
        fecha_str = t.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_hora_entrega else "-"
        entregado_str = "Sí" if t.entregado else "No"
        recolector_str = t.nombre_recolector if t.nombre_recolector else t.nombre_alumno

        ws.append([
            t.numero_boleto,
            t.codigo_alumno,
            t.nombre_alumno,
            t.carrera,
            t.ciclo,
            recolector_str,
            t.estado.upper(),
            t.monto_total,
            t.monto_pagado,
            t.monto_pendiente,
            t.metodo_pago.upper(),
            entregado_str,
            fecha_str
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.column in [1, 2, 5, 7, 11, 12]:
                cell.alignment = Alignment(horizontal="center")
            elif cell.column in [8, 9, 10]:
                cell.number_format = 'S/ #,##0.00'

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_boletos_{evento.nombre.replace(' ', '_')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/{evento_id}/exportar/csv")
def exportar_tickets_csv(
    evento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Exporta la lista completa de boletos vendidos/separados a un archivo CSV.
    Requiere autenticación.
    """
    evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento no encontrado"
        )

    tickets = db.query(Ticket).filter(Ticket.evento_id == evento_id).order_by(Ticket.numero_boleto.asc()).all()

    output = io.StringIO()
    output.write('\ufeff') # UTF-8 BOM
    writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "Nº Boleto", "Código / DNI", "Nombre Comprador", "Carrera", "Ciclo",
        "Persona que Recoge", "Estado", "Monto Total (S/)", "Monto Pagado (S/)",
        "Monto Pendiente (S/)", "Método de Pago", "Entregado", "Fecha de Entrega"
    ])

    for t in tickets:
        fecha_str = t.fecha_hora_entrega.strftime("%Y-%m-%d %H:%M:%S") if t.fecha_hora_entrega else ""
        entregado_str = "Sí" if t.entregado else "No"
        recolector_str = t.nombre_recolector if t.nombre_recolector else t.nombre_alumno

        writer.writerow([
            t.numero_boleto,
            t.codigo_alumno,
            t.nombre_alumno,
            t.carrera,
            t.ciclo,
            recolector_str,
            t.estado.upper(),
            f"{t.monto_total:.2f}",
            f"{t.monto_pagado:.2f}",
            f"{t.monto_pendiente:.2f}",
            t.metodo_pago.upper(),
            entregado_str,
            fecha_str
        ])

    filename = f"reporte_boletos_{evento.nombre.replace(' ', '_')}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("", response_model=EventoOut, status_code=status.HTTP_201_CREATED)
def crear_evento(
    evento_in: EventoCreate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    db_evento = Evento(
        nombre=evento_in.nombre,
        fecha=evento_in.fecha,
        activo=evento_in.activo
    )
    db.add(db_evento)
    db.commit()
    db.refresh(db_evento)
    return db_evento

@router.get("", response_model=List[EventoOut])
def listar_eventos(
    activo: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(Evento)
    if activo is not None:
        query = query.filter(Evento.activo == activo)
    return query.all()

@router.get("/{evento_id}", response_model=EventoOut)
def obtener_evento(
    evento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    db_evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not db_evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento no encontrado"
        )
    return db_evento

@router.put("/{evento_id}", response_model=EventoOut)
def actualizar_evento(
    evento_id: int,
    evento_in: EventoUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    db_evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not db_evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento no encontrado"
        )
    
    update_data = evento_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_evento, field, value)

    db.commit()
    db.refresh(db_evento)
    return db_evento

@router.delete("/{evento_id}", status_code=status.HTTP_204_NO_CONTENT)
def eliminar_evento(
    evento_id: int,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    db_evento = db.query(Evento).filter(Evento.id == evento_id).first()
    if not db_evento:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento no encontrado"
        )
    db.delete(db_evento)
    db.commit()
    return None
