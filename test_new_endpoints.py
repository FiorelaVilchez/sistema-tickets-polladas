import os
import sys
from fastapi.testclient import TestClient
import openpyxl
import io

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app.main import app
from seed import seed_database

def run_tests():
    print("=== Inicializando BD de prueba ===")
    seed_database()
    client = TestClient(app)

    print("\n=== Login para obtener Token JWT ===")
    resp_login = client.post("/auth/login", data={"username": "admin", "password": "admin123"})
    assert resp_login.status_code == 200
    token = resp_login.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Obtener ID de un evento existente
    resp_eventos = client.get("/eventos", headers=headers)
    assert resp_eventos.status_code == 200
    evento_id = resp_eventos.json()[0]["id"]

    print("\n--- TEST 1: Registrar Venta / Venta Adicional ---")
    venta_payload = {
        "evento_id": evento_id,
        "nombre_alumno": "Carlos Mendoza",
        "codigo_alumno": "202400111",
        "estado": "separado"
    }
    r1 = client.post("/tickets/registrar-venta", json=venta_payload, headers=headers)
    assert r1.status_code == 201, f"Error al registrar venta: {r1.text}"
    ticket1 = r1.json()
    codigo1 = ticket1["codigo_unico"]
    assert ticket1["estado"] == "separado"
    print(f"[OK] Ticket 1 registrado: {codigo1} (separado)")

    # Venta de ticket adicional
    venta_payload2 = {
        "evento_id": evento_id,
        "nombre_alumno": "Carlos Mendoza",
        "codigo_alumno": "202400111",
        "estado": "vendido"
    }
    r2 = client.post("/tickets/registrar-venta", json=venta_payload2, headers=headers)
    assert r2.status_code == 201
    ticket2 = r2.json()
    codigo2 = ticket2["codigo_unico"]
    print(f"[OK] Ticket Adicional registrado: {codigo2} (vendido)")

    print("\n--- TEST 2: Buscar Ticket ---")
    # Buscar por codigo_alumno
    rb1 = client.get(f"/tickets/buscar?codigo_alumno=202400111", headers=headers)
    assert rb1.status_code == 200
    tickets_alumno = rb1.json()
    assert len(tickets_alumno) >= 2
    print(f"[OK] Se encontraron {len(tickets_alumno)} tickets para el alumno 202400111.")

    # Buscar por codigo_unico
    rb2 = client.get(f"/tickets/buscar?codigo_unico={codigo1}", headers=headers)
    assert rb2.status_code == 200
    assert len(rb2.json()) == 1
    assert rb2.json()[0]["codigo_unico"] == codigo1
    print(f"[OK] Búsqueda por codigo_unico {codigo1} exitosa.")

    print("\n--- TEST 3 & 4: Confirmar Entrega y Manejo de Duplicados ---")
    # Primera entrega (debe ser exitosa)
    re1 = client.post("/tickets/confirmar-entrega", json={"codigo_unico": codigo1}, headers=headers)
    assert re1.status_code == 200
    data_e1 = re1.json()
    assert data_e1["entregado"] is True
    assert data_e1["fecha_hora_entrega"] is not None
    print(f"[OK] Primera entrega confirmada para {codigo1} a las {data_e1['fecha_hora_entrega']}")

    # Segunda entrega (debe fallar con error 400 explícito)
    re2 = client.post("/tickets/confirmar-entrega", json={"codigo_unico": codigo1}, headers=headers)
    assert re2.status_code == 400
    error_msg = re2.json()["detail"]
    assert "ya fue entregado previamente" in error_msg
    print(f"[OK] Intento de entrega duplicada capturado correctamente: '{error_msg}'")

    print("\n--- TEST 5: KPIs del Evento ---")
    rk = client.get(f"/eventos/{evento_id}/kpis", headers=headers)
    assert rk.status_code == 200
    kpis = rk.json()
    print(f"[OK] KPIs del evento '{kpis['nombre_evento']}':")
    print(f"     Total Tickets: {kpis['total_tickets']}")
    print(f"     Vendidos: {kpis['vendidos']}")
    print(f"     Separados: {kpis['separados']}")
    print(f"     No Vendidos: {kpis['no_vendidos']}")
    print(f"     Entregados: {kpis['entregados']}")
    assert kpis["total_tickets"] > 0
    assert kpis["entregados"] >= 1

    print("\n--- TEST 6: Exportación a Excel y CSV ---")
    # Excel
    re_excel = client.get(f"/eventos/{evento_id}/exportar/excel", headers=headers)
    assert re_excel.status_code == 200
    assert "spreadsheetml" in re_excel.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(re_excel.content))
    ws = wb.active
    assert ws.max_row >= 2 # Encabezado + filas
    print(f"[OK] Excel generado con éxito ({len(re_excel.content)} bytes, {ws.max_row} filas en total).")

    # CSV
    re_csv = client.get(f"/eventos/{evento_id}/exportar/csv", headers=headers)
    assert re_csv.status_code == 200
    assert "text/csv" in re_csv.headers["content-type"]
    csv_text = re_csv.content.decode('utf-8')
    assert "Código Único" in csv_text
    assert "Carlos Mendoza" in csv_text
    print(f"[OK] CSV generado con éxito ({len(re_csv.content)} bytes).")

    print("\n=======================================================")
    print("  TODOS LOS NUEVOS REQUERIMIENTOS FUERON VERIFICADOS  ")
    print("=======================================================")

if __name__ == "__main__":
    run_tests()
